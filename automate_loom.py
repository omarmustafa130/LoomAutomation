from playwright.sync_api import sync_playwright
import time
import tempfile
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from openpyxl import Workbook, load_workbook
import threading
import queue
import json
from pathlib import Path
import sys
import shutil



PAUSE_FLAG = False
excel_lock = threading.Lock()
# Configuration constants
CONFIG_FILE = "loom_config.json"
LOOM_COOKIES_FILE = "loom_cookies.json"
TEMPORARY_DOWNLOAD_DIR = 'downloaded_videos'
os.makedirs(TEMPORARY_DOWNLOAD_DIR, exist_ok=True)

# Set browser path for frozen executable
if getattr(sys, 'frozen', False):
    exe_path = os.path.dirname(sys.executable)
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = os.path.join(exe_path, 'browsers')

# Helper functions
def load_config():
    config = {'folder_id': '', 'service_file': '', 'space': ''}
    try:
        if Path(CONFIG_FILE).exists():
            with open(CONFIG_FILE, 'r') as f:
                config.update(json.load(f))
    except Exception as e:
        print(f"Error loading config: {e}")
    return config

def save_config():
    config = {
        'folder_id': folder_id_entry.get(),
        'service_file': service_file_entry.get(),
        'space': space_entry.get()
    }
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f)


def append_to_excel(video_title, video_url, embed_code):
    file_name = "uploaded_videos.xlsx"
    sheet_name = "Videos"
    
    with excel_lock:
        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Video Title", "URL", "Embed Code"])
        else:
            ws = wb[sheet_name]
        
        ws.append([video_title, video_url, embed_code])
        wb.save(file_name)

def update_excel_embed_code(video_url, new_embed_code):
    file_name = "uploaded_videos.xlsx"
    with excel_lock:
        try:
            wb = load_workbook(file_name)
            ws = wb["Videos"]
            for row in ws.iter_rows(min_row=2):
                if row[1].value == video_url:
                    row[2].value = new_embed_code
            wb.save(file_name)
            return True
        except Exception as e:
            print(f"Error updating Excel: {e}")
            return False



def login_and_save_cookies(progress_queue):
    with sync_playwright() as p:
        temp_profile_dir = tempfile.mkdtemp()
        print(f"[LOGIN] Using profile at: {temp_profile_dir}")
        context = p.chromium.launch_persistent_context(
            temp_profile_dir,
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ]
        )
        page = context.new_page()
        page.evaluate("() => { delete window.navigator.webdriver; }")
        page.goto("https://www.loom.com/looms/videos")
        progress_queue.put(("info", "Login Required", "Please login to Loom in the opened browser.\nPress OK here when done."))
        # Wait for user to log in (this could be improved, e.g., with a manual signal)
        time.sleep(10)
        cookies = context.cookies()
        print("[LOGIN] Cookies extracted. Saving to file loom_cookies.json.")
        with open(LOOM_COOKIES_FILE, "w") as f:
            json.dump(cookies, f, indent=2)
        context.close()
        progress_queue.put(("info", "Login Complete", "Cookies have been saved. You can close the browser now if it's still open."))


def logout():
    """Delete config and cookies file, then clear GUI entries."""
    confirm = messagebox.askyesno(
        "Confirm Logout",
        "This will delete your saved Loom cookies and config file. Continue?"
    )
    if not confirm:
        return  # User cancelled

    # Delete loom_cookies.json and loom_config.json if they exist
    try:
        if os.path.exists(LOOM_COOKIES_FILE):
            os.remove(LOOM_COOKIES_FILE)
        if os.path.exists(CONFIG_FILE):
            os.remove(CONFIG_FILE)
    except Exception as e:
        messagebox.showerror("Error", f"Error removing files: {e}")
        return

    # Optionally clear text fields in the GUI
    folder_id_entry.delete(0, tk.END)
    service_file_entry.delete(0, tk.END)
    space_entry.delete(0, tk.END)

    messagebox.showinfo("Logged Out", "Cookies and config file have been removed.")




# Video processing functions
def extract_embed_code(page, progress_queue, title):
    try:
        page.wait_for_selector('button[data-testid="share-modal-button"]', timeout=20000).click()
        time.sleep(10)
        dialog = page.locator("dialog.css-1gw7q29[role='dialog']")
        dialog.locator("button.menu_shareTab_3H-", has_text="Embed").click()
        
        try:
            page.wait_for_selector('img[alt="Video thumbnail"]', timeout=5000)
        except:
            dialog = page.locator("dialog.css-1gw7q29[role='dialog']")
            dialog.locator("button.menu_shareTab_3H-", has_text="Embed").click()
        copy_btn = page.wait_for_selector('button.css-ask8uh:has-text("Copy embed code")', timeout=20000)
        copy_btn.click()
        embed_code = page.evaluate("navigator.clipboard.readText()")
        progress_queue.put(("embed_success", title, page.url, embed_code))
        return embed_code
    except Exception as e:
        progress_queue.put(("embed_error", page.url, str(e)))
        return None
    


def process_video_url(url, progress_queue, title):
    if not os.path.exists(LOOM_COOKIES_FILE):
        return None

    with open(LOOM_COOKIES_FILE, "r") as f:
        stored_cookies = json.load(f)
        
    with sync_playwright() as p:
        temp_profile_dir = tempfile.mkdtemp()
        context = p.chromium.launch_persistent_context(
            temp_profile_dir,
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            ]
        )
        context.grant_permissions(["clipboard-read", "clipboard-write"], origin="https://www.loom.com")
        context.add_cookies(stored_cookies)

        page = context.new_page()
        try:
            page.goto(url)
            # Possibly track how long it's taking...
            # You can add your own "stuck" logic similarly.

            time.sleep(5)
            embed_code = extract_embed_code(page, progress_queue, title)
            return embed_code
        except Exception as e:
            progress_queue.put(("embed_error", url, str(e)))
            return None
        finally:
            context.close()
            shutil.rmtree(temp_profile_dir, ignore_errors=True)

def sync_videos(progress_queue):
    progress_queue.put(("status", "Starting video synchronization..."))
    if not os.path.exists(LOOM_COOKIES_FILE):
        progress_queue.put(("error", "No cookies found. Please login first."))
        return

    space_url = space_entry.get().strip()
    if not space_url:
        progress_queue.put(("error", "Please provide a Space URL"))
        return

    with open(LOOM_COOKIES_FILE, "r") as f:
        stored_cookies = json.load(f)

    with sync_playwright() as p:
        temp_profile_dir = tempfile.mkdtemp()
        context = p.chromium.launch_persistent_context(
            temp_profile_dir,
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
            ]
        )
        context.add_cookies(stored_cookies)
        page = context.new_page()

        try:
            progress_queue.put(("status", "Loading Loom space content..."))
            page.goto(space_url)
            page.wait_for_selector('article[data-videoid]', timeout=30000)

            prev_count = 0
            max_consecutive_no_new = 6
            consecutive_no_new = 0

            while consecutive_no_new < max_consecutive_no_new:
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(8)  # Wait for content to load
                current_videos = page.query_selector_all('article[data-videoid]')
                current_count = len(current_videos)
                progress_queue.put(("status", f"Loaded {current_count} videos..."))
                if current_count > prev_count:
                    prev_count = current_count
                    consecutive_no_new = 0  # Reset if new videos are found
                else:
                    consecutive_no_new += 1  # Increment if no new videos

            # After scrolling, get all videos
            videos = page.query_selector_all('article[data-videoid]')
            progress_queue.put(("status", f"Found {len(videos)} videos in space"))

            existing_urls = set()
            if os.path.exists("uploaded_videos.xlsx"):
                wb = load_workbook("uploaded_videos.xlsx")
                ws = wb.active
                existing_urls = {row[1].value for row in ws.iter_rows(min_row=2)}

            new_entries = 0
            for i, video in enumerate(videos, 1):
                link = video.query_selector('a.video-card_videoCardLink_37D')
                if not link:
                    print(f"Video {i}: No link found.")
                    continue

                url = link.get_attribute('href')
                title_attr = link.get_attribute('aria-label')
                if not url or not title_attr:
                    print(f"Video {i}: Missing URL or title.")
                    continue

                title = title_attr.replace('Open video: ', '')
                print(f"Video {i}: {title} - {url}")

                if url not in existing_urls:
                    append_to_excel(title, url, "")
                    new_entries += 1
                else:
                    print(f"Video {i}: Duplicate URL.")


            progress_queue.put(("status", f"Sync complete! Added {new_entries} new videos"))
            
        except Exception as e:
            progress_queue.put(("error", f"Sync failed: {str(e)}"))
        finally:
            context.close()
            shutil.rmtree(temp_profile_dir, ignore_errors=True)
            
def generate_embed_codes(progress_queue):
    max_retries = 3
    
    try:
        with excel_lock:
            wb = load_workbook("uploaded_videos.xlsx")
            ws = wb["Videos"]
            rows_to_process = []
            for row in ws.iter_rows(min_row=2):
                title = row[0].value or ""
                url   = row[1].value or ""
                code  = row[2].value or ""
                if not code or "Couldn't Extract" in code:
                    rows_to_process.append((title, url))
        
        if not rows_to_process:
            progress_queue.put(("complete", "All embed codes already present"))
            return
        
        progress_queue.put(("total_embeds", len(rows_to_process)))
        
        for i, (title, url) in enumerate(rows_to_process, start=1):
            progress_queue.put(("current_embed", i, url))
            attempts_left = max_retries
            embed_code = None
            
            while attempts_left > 0:
                try:
                    embed_code = process_video_url(url, progress_queue, title)
                    if embed_code:
                        break
                except Exception as e:
                    print(f"Embed extraction error for {url}: {e}")
                attempts_left -= 1
                if attempts_left > 0:
                    print(f"Retrying embed extraction for {url}...")
            
            if embed_code:
                update_excel_embed_code(url, embed_code)
            else:
                embed_code = "Couldn't Extract - Retry Failed"
                update_excel_embed_code(url, embed_code)
    
    except Exception as e:
        progress_queue.put(("embed_error", "General Error", str(e)))
    finally:
        progress_queue.put(("complete", "Embed code generation completed"))



# GUI functions
def start_generate_embeds():
    if not os.path.exists("uploaded_videos.xlsx"):
        messagebox.showerror("Error", "No Excel file found")
        return
        
    # Create our progress queue for communication
    progress_queue = queue.Queue()

    # 1) Clear the TreeView right away
    progress_queue.put(("clear_tree", None))

    # 2) Launch thread to do the embed extraction
    threading.Thread(target=generate_embed_codes, args=(progress_queue,)).start()

    # 3) Keep checking queue
    root.after(100, lambda: check_progress_queue(progress_queue))


def pause_upload():
    global PAUSE_FLAG
    PAUSE_FLAG = True
    messagebox.showinfo("Paused", "Upload process will paused now")

def start_download_and_upload():
    """Download all videos, then immediately upload them (in one go)."""
    folder_id = folder_id_entry.get().strip()
    service_file = service_file_entry.get().strip()
    if not folder_id or not service_file:
        messagebox.showerror("Error", "Please provide both Folder ID and Service JSON file.")
        return

    # Create a single queue to feed progress to check_progress_queue
    progress_queue = queue.Queue()

    def run_both():
        # Call download_videos logic
        download_videos(folder_id, service_file, progress_queue)

        # Then call upload_videos logic
        upload_videos(progress_queue)

        # Signal "complete"
        progress_queue.put(("complete", None))

    # Run in background so GUI doesn't freeze
    threading.Thread(target=run_both).start()
    # Keep checking progress_queue
    root.after(100, lambda: check_progress_queue(progress_queue))

# Function to fetch video files from Google Drive
# Update the get_gdrive_videos function
def get_gdrive_videos(drive_service, folder_id):
    """Fetch video files from Google Drive folder"""
    # List of common video MIME types
    video_mime_types = [
        'video/',  # This covers all video types generically
        'application/vnd.google-apps.video'  # Google Drive native video format
    ]
    
    # Build the query with explicit MIME type checks
    query = f"""
        '{folder_id}' in parents and (
            mimeType contains 'video/' or
            mimeType = 'application/vnd.google-apps.video'
        )
    """
    
    results = drive_service.files().list(
        q=query.strip(),
        fields="files(id, name, mimeType, fileExtension)",
        pageSize=1000
    ).execute()
    
    # Additional validation for returned files
    valid_files = []
    for file in results.get('files', []):
        # Check if it's a Google Drive native video file
        if file['mimeType'] == 'application/vnd.google-apps.video':
            valid_files.append(file)
            continue
            
        # Check for standard video files
        if (file['mimeType'].startswith('video/') and 
            file.get('fileExtension', '').lower() in {'mp4', 'mov', 'avi', 'mkv', 'flv', 'wmv'}):
            valid_files.append(file)
    
    return valid_files

def download_video(drive_service, file_id, file_name, on_progress=None):
    """Download a video with progress reporting."""
    file_path = os.path.join(TEMPORARY_DOWNLOAD_DIR, file_name)
    request = drive_service.files().get_media(fileId=file_id)
    with open(file_path, 'wb') as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            if status and on_progress:
                # Report progress percentage for this file
                percent = int(status.progress() * 100)
                on_progress(percent)
    return file_path

def pause_upload():
    global PAUSE_FLAG
    PAUSE_FLAG = True
    messagebox.showinfo("Paused", "Upload process will pause now")


def download_videos(folder_id, service_file, progress_queue):
    """Download each video with its own progress bar reset per video."""
    credentials = service_account.Credentials.from_service_account_file(
        service_file,
        scopes=['https://www.googleapis.com/auth/drive.readonly']
    )
    drive_service = build('drive', 'v3', credentials=credentials)
    videos = get_gdrive_videos(drive_service, folder_id)
    
    total_videos = len(videos)

    for i, video in enumerate(videos):
        # Notify start of new video download
        progress_queue.put(("download", f"Starting download: {video['name']}", 0))

        def progress_callback(percent):
            """Update UI with per-video progress."""
            progress_queue.put(("download", f"Downloading: {video['name']} ({percent}%)", percent))

        # Download with progress tracking
        download_video(drive_service, video['id'], video['name'], progress_callback)

        # Notify completion of the video download
        progress_queue.put(("download", f"Downloaded: {video['name']} ✔", 100))
        time.sleep(1)  # Small delay for visual clarity
        progress_queue.put(("download", f"", 0))  # Reset progress bar for next video

    # Update UI with downloaded files
    progress_queue.put(("populate_listbox", os.listdir(TEMPORARY_DOWNLOAD_DIR))) 
    save_config()


def upload_videos(progress_queue):
    global PAUSE_FLAG
    PAUSE_FLAG = False

    max_retries = 3
    upload_timeout_seconds = 180
    processing_timeout = 180000
    time_between_checks = 5
    stuck_threshold = 60
    max_upload_time = 600

    save_config()
    files_to_upload = os.listdir(TEMPORARY_DOWNLOAD_DIR)
    progress_queue.put(("populate_listbox", files_to_upload))

    if not files_to_upload:
        progress_queue.put(("complete", None))
        return
    
    if not os.path.exists(LOOM_COOKIES_FILE):
        progress_queue.put(("error", "No loom_cookies.json found. Please log in first."))
        progress_queue.put(("Not logged in", None))
        return

    try:
        with open(LOOM_COOKIES_FILE, "r") as f:
            stored_cookies = json.load(f)
    except Exception as e:
        progress_queue.put(("error", f"Could not load cookies file: {e}"))
        progress_queue.put(("Could not log in", None))
        return

    with sync_playwright() as p:
        i = 0
        while i < len(files_to_upload):
            filename = files_to_upload[i]
            if PAUSE_FLAG:
                progress_queue.put(("pausing", None))
                break

            file_path = os.path.join(TEMPORARY_DOWNLOAD_DIR, filename)
            if not os.path.isfile(file_path):
                i += 1
                continue

            file_size = os.path.getsize(file_path)
            attempts_left = max_retries
            success = False

            while attempts_left > 0 and not success:
                temp_profile_dir = tempfile.mkdtemp()
                try:
                    # Launch browser context
                    context = p.chromium.launch_persistent_context(
                        temp_profile_dir,
                        headless=True,
                        viewport={"width": 1280, "height": 720},
                        args=[
                            "--disable-blink-features=AutomationControlled",
                            "--no-sandbox",
                            "--disable-dev-shm-usage",
                            "--disable-infobars",
                            "--window-size=1280,720",
                            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
                        ]
                    )
                    context.add_cookies(stored_cookies)
                    context.grant_permissions(["clipboard-read", "clipboard-write"], origin="https://www.loom.com")

                    page = context.new_page()
                    page.evaluate("() => { delete window.navigator.webdriver; }")

                    progress_queue.put(("upload", f"Opening Loom workspace for {filename}...", 0))
                    space_url = space_entry.get().strip()
                    page.goto(space_url)
                    time.sleep(8)

                    # Click "Add video" button with timeout
                    try:
                        add_video_button = page.wait_for_selector('button:has-text("Add video")', timeout=60000)
                        add_video_button.click(force=True)
                    except TimeoutError:
                        print(f"Timeout waiting for 'Add video' button for {filename}")
                        raise
                    time.sleep(3)
                    progress_queue.put(("upload", "Initiating upload...", 0))

                    # Select "Upload a video" option
                    try:
                        upload_option = page.wait_for_selector('li[role="option"]:has-text("Upload a video")', timeout=60000)
                        upload_option.click(force=True)
                    except TimeoutError:
                        print(f"Timeout waiting for upload option for {filename}")
                        raise
                    time.sleep(3)

                    # Trigger file chooser and upload
                    with page.expect_file_chooser() as fc_info:
                        page.keyboard.press(" ")
                    file_chooser = fc_info.value
                    file_chooser.set_files(file_path)
                    time.sleep(2)
                    try:
                        page.wait_for_selector("text=Upload 1 file", timeout=60000)
                        upload_button = page.wait_for_selector("button.uppy-StatusBar-actionBtn--upload", timeout=60000)
                        upload_button.click(force=True)
                        time.sleep(2)
                    except TimeoutError:
                        print(f"Timeout initiating upload for {filename}")
                        raise

                    # Upload progress monitoring
                    previous_percentage = 0
                    previous_time = time.time()
                    last_progress_update = time.time()
                    upload_start_time = time.time()

                    while True:
                        if PAUSE_FLAG:
                            context.close()
                            shutil.rmtree(temp_profile_dir, ignore_errors=True)
                            progress_queue.put(("pausing", None))
                            return

                        # Check total upload time
                        if time.time() - upload_start_time > max_upload_time:
                            raise TimeoutError(f"Upload for {filename} exceeded maximum time of {max_upload_time} seconds")

                        # Check upload status
                        try:
                            status_element = page.wait_for_selector(".uppy-StatusBar-statusPrimary", timeout=60000)
                            status_text = status_element.inner_text().strip()
                        except TimeoutError:
                            print(f"Timeout waiting for status element for {filename}")
                            raise

                        if "Uploading: " in status_text:
                            current_percentage = int(status_text.split(": ")[1].replace("%", ""))
                            if current_percentage != previous_percentage:
                                now = time.time()
                                delta_percent = current_percentage - previous_percentage
                                bytes_uploaded_now = (delta_percent / 100.0) * file_size
                                delta_time = now - previous_time or 0.1
                                speed_mbs = (bytes_uploaded_now / 1_000_000) / delta_time

                                previous_percentage = current_percentage
                                previous_time = now
                                last_progress_update = now

                                progress_queue.put((
                                    "upload",
                                    f"Uploading {filename}: {current_percentage}% ({speed_mbs:.2f} MB/s)",
                                    current_percentage
                                ))
                            elif time.time() - last_progress_update > stuck_threshold:
                                raise TimeoutError(f"Upload stuck at {current_percentage}% for over {stuck_threshold} seconds")

                        elif "Complete" in status_text:
                            progress_queue.put(("upload", f"{filename}: 100% Complete", 100))
                            break
                        else:
                            # If status is neither "Uploading" nor "Complete", check if stuck
                            if time.time() - last_progress_update > stuck_threshold:
                                raise TimeoutError(f"Upload in unknown state '{status_text}' for over {stuck_threshold} seconds")

                        time.sleep(time_between_checks)
                    # Extract URL
                    progress_queue.put(("upload", "Finished uploading. Extracting URL..", 0))
                    try:
                        time.sleep(5)
                        
                        page.wait_for_selector(".uppy-Dashboard-Item.is-complete", timeout=processing_timeout)
                        video_link_element = page.wait_for_selector(
                            ".uppy-Dashboard-Item.is-complete .uppy-Dashboard-Item-previewLink",
                            timeout=processing_timeout
                        )
                        video_url = video_link_element.get_attribute("href")
                        
                    except TimeoutError as e:
                        append_to_excel(filename, "", "")
                        print(f"Timeout extracting URL for {filename}: {e}")
                        raise
                    
                    try:
                        # Check if the video URL is valid
                        page.goto(video_url)
                        page.wait_for_selector('button[data-testid="share-modal-button"]', timeout=20000).click()
                        time.sleep(10)
                        dialog = page.locator("dialog.css-1gw7q29[role='dialog']")
                        dialog.locator("button.menu_shareTab_3H-", has_text="Embed").click()
                        try:
                            page.wait_for_selector('img[alt="Video thumbnail"]', timeout=5000)
                        except:
                            dialog = page.locator("dialog.css-1gw7q29[role='dialog']")
                            dialog.locator("button.menu_shareTab_3H-", has_text="Embed").click()
                        copy_btn = page.wait_for_selector('button.css-ask8uh:has-text("Copy embed code")', timeout=20000)
                        copy_btn.click()
                        embed_code = page.evaluate("navigator.clipboard.readText()")
                        
                    except:
                        embed_code = ""
                        
                    append_to_excel(filename, video_url, embed_code)
                    progress_queue.put(("add_video", filename, video_url, embed_code))

                    os.remove(file_path)
                    progress_queue.put(("remove_file", filename))

                    context.close()
                    shutil.rmtree(temp_profile_dir, ignore_errors=True)
                    success = True

                except TimeoutError as toe:
                    attempts_left -= 1
                    print(f"TimeoutError for {filename}: {toe}. Attempts left: {attempts_left}")
                    context.close()
                    shutil.rmtree(temp_profile_dir, ignore_errors=True)
                    if attempts_left == 0:
                        progress_queue.put(("warning", f"Skipped {filename} after {max_retries} failed attempts due to timeout."))
                        progress_queue.put(("upload", f"Skipped {filename} due to repeated timeouts", 0))
                        break

                except Exception as e:
                    attempts_left -= 1
                    print(f"Unexpected error uploading {filename}: {e}")
                    context.close()
                    shutil.rmtree(temp_profile_dir, ignore_errors=True)
                    if attempts_left == 0:
                        progress_queue.put(("error", f"Skipped {filename} after {max_retries} failed attempts due to error: {e}"))
                        progress_queue.put(("upload", f"Skipped {filename} due to repeated errors", 0))
                        break

            i += 1

        progress_queue.put(("complete", None))

# Browse button for service file
def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    if file_path:
        service_file_entry.delete(0, tk.END)
        service_file_entry.insert(0, file_path)


# <--- NEW CODE: A "Login" button
def start_login():
    progress_queue = queue.Queue()
    def run_login():
        login_and_save_cookies(progress_queue)
    t = threading.Thread(target=run_login)
    t.start()
    root.after(100, lambda: check_progress_queue(progress_queue))
    
# Download Videos button
def start_download():
    folder_id = folder_id_entry.get().strip()
    service_file = service_file_entry.get().strip()
    if not folder_id or not service_file:
        messagebox.showerror("Error", "Please provide both Folder ID and Service JSON file.")
        return
    progress_queue = queue.Queue()
    download_thread = threading.Thread(target=download_videos, args=(folder_id, service_file, progress_queue))
    download_thread.start()
    root.after(100, lambda: check_progress_queue(progress_queue))

# Rename Selected button
def rename_selected():
    selection = upload_listbox.curselection()
    if not selection:
        messagebox.showwarning("No Selection", "Please select a video to rename.")
        return
    index = selection[0]
    old_name = upload_listbox.get(index)
    old_base, old_ext = os.path.splitext(old_name)
    
    # Fixed dialog without width parameter
    new_base = simpledialog.askstring(
        "Rename Video", 
        f"Enter new name (extension {old_ext} will be kept):\n",
        initialvalue=old_base,
        parent=root
    )
    
    if new_base and new_base != old_base:
        new_name = f"{new_base}{old_ext}"
        old_path = os.path.join(TEMPORARY_DOWNLOAD_DIR, old_name)
        new_path = os.path.join(TEMPORARY_DOWNLOAD_DIR, new_name)
        try:
            os.rename(old_path, new_path)
            upload_listbox.delete(index)
            upload_listbox.insert(index, new_name)
        except Exception as e:
            messagebox.showerror("Rename Error", f"Failed to rename file: {e}")

# Upload Videos button
def start_upload():
    if not os.listdir(TEMPORARY_DOWNLOAD_DIR):
        messagebox.showwarning("No Videos", "No videos to upload. Please download videos first.")
        return
    progress_queue = queue.Queue()
    upload_thread = threading.Thread(target=upload_videos, args=(progress_queue,))
    upload_thread.start()
    root.after(100, lambda: check_progress_queue(progress_queue))

def check_progress_queue(progress_queue):
    try:
        while True:
            item = progress_queue.get_nowait()
            if item[0] == "populate_listbox":
                upload_listbox.delete(0, tk.END)
                for fname in item[1]:
                    upload_listbox.insert(tk.END, fname)
            elif item[0] == "add_video":
                tree.insert("", "end", values=(item[1], item[2], item[3]))
            elif item[0] in ("download", "upload"):
                _, text, value = item
                progress_label.config(text=f"{text} ({value}%)")
                progress_bar['value'] = value
            elif item[0] == "status":
                progress_label.config(text=item[1])
            elif item[0] == "remove_file":
                filename = item[1]
                all_items = upload_listbox.get(0, tk.END)
                if filename in all_items:
                    idx = all_items.index(filename)
                    upload_listbox.delete(idx)
            elif item[0] == "complete":
                msg = item[1] if item[1] else "Operation Complete"
                messagebox.showinfo("Complete", msg)
                progress_bar['value'] = 0
                progress_label.config(text="Operation Complete")
            elif item[0] == "total_embeds":
                progress_bar['maximum'] = item[1]
            elif item[0] == "current_embed":
                progress_label.config(text=f"Processing {item[1]}/{progress_bar['maximum']}: {item[2]}")
                progress_bar['value'] = item[1]
            elif item[0] == "clear_tree":
                for row_id in tree.get_children():
                    tree.delete(row_id)
            elif item[0] == "embed_success":
                title = item[1]
                url   = item[2]
                code  = item[3]
                tree.insert("", "end", values=(title, url, code))
            elif item[0] == "embed_error":
                progress_queue.put("Error", f"Failed to process {item[1]}: {item[2]}")
            elif item[0] == "error":
                progress_queue.put("Error", item[1])
            elif item[0] == "warning":
                progress_queue.put("Warning", item[1])
            elif item[0] == "info":
                progress_queue.put(item[1], item[2])
            elif item[0] == "pausing":
                progress_label.config(text="Upload paused")
                progress_bar['value'] = 0
                messagebox.showinfo("Paused", "Upload has been paused.")
            root.update_idletasks()
    except queue.Empty:
        pass
    root.after(100, lambda: check_progress_queue(progress_queue))

def start_sync():
    progress_queue = queue.Queue()
    threading.Thread(target=sync_videos, args=(progress_queue,)).start()
    root.after(100, lambda: check_progress_queue(progress_queue))


root = tk.Tk()
root.title("Loom Video Uploader")
root.geometry("900x650")  # Wider window
root.configure(bg='#f0f0f0')
root.resizable(False, False)

# Style configuration
style = ttk.Style()
style.theme_use('clam')
style.configure(".", background='#f0f0f0', foreground='black')
style.configure("TButton", background='#e1e1e1')
style.configure("Horizontal.TProgressbar", background="green", troughcolor='#d0d0d0')
style.configure("Treeview", background='white', fieldbackground='white')
style.map("Treeview", background=[('selected', '#347083')])

# Frames
input_frame = ttk.Frame(root, padding=10)
input_frame.pack(pady=5, fill='x')

# Load config
config = load_config()

# Input fields
ttk.Label(input_frame, text="Google Drive Folder ID:").grid(row=0, column=0, padx=5, sticky='w')
folder_id_entry = ttk.Entry(input_frame, width=60)
folder_id_entry.insert(0, config['folder_id'])
folder_id_entry.grid(row=0, column=1, padx=5, sticky='ew')

ttk.Label(input_frame, text="Service Account JSON File:").grid(row=1, column=0, padx=5, sticky='w')
service_file_entry = ttk.Entry(input_frame, width=60)
service_file_entry.insert(0, config['service_file'])
service_file_entry.grid(row=1, column=1, padx=5, sticky='ew')

ttk.Label(input_frame, text="Space URL:").grid(row=2, column=0, padx=5, sticky='w')
space_entry = ttk.Entry(input_frame, width=102)
space_entry.insert(0, config['space'])
space_entry.grid(row=2, column=1, padx=5, sticky='ew')

browse_button = ttk.Button(input_frame, text="Browse", command=browse_file)
browse_button.grid(row=1, column=2, padx=5)

# Existing buttons
button_frame = ttk.Frame(root)
button_frame.pack(pady=5)

login_button = ttk.Button(button_frame, text="Login", command=start_login)
logout_button = ttk.Button(button_frame, text="Logout", command=logout)
download_button = ttk.Button(button_frame, text="Download", command=start_download)
download_upload_button = ttk.Button(button_frame, text="Download & Upload", command=start_download_and_upload)
rename_button = ttk.Button(button_frame, text="Rename", command=rename_selected)
upload_button = ttk.Button(button_frame, text="Upload", command=start_upload)
pause_button = ttk.Button(button_frame, text="Pause", command=pause_upload)
generate_button = ttk.Button(button_frame, text="Generate Embeds", command=start_generate_embeds)
sync_button = ttk.Button(button_frame, text="Sync", command=lambda: start_sync())

buttons = [login_button, logout_button, download_button, download_upload_button, rename_button, upload_button, pause_button, generate_button,sync_button]
for btn in buttons:
    btn.pack(side='left', padx=5)

# List and progress elements remain the same
list_frame = ttk.Frame(root, padding=10)
list_frame.pack(pady=5, fill='both', expand=True)

upload_label = ttk.Label(list_frame, text="Videos to Upload")
upload_label.pack(anchor='w')

list_container = ttk.Frame(list_frame)
list_container.pack(fill='both', expand=True)

scrollbar = ttk.Scrollbar(list_container)
upload_listbox = tk.Listbox(list_container, selectmode="single", bg='white', yscrollcommand=scrollbar.set)
scrollbar.config(command=upload_listbox.yview)
upload_listbox.pack(side='left', fill='both', expand=True)
scrollbar.pack(side='right', fill='y')

# Progress elements
progress_frame = ttk.Frame(root, padding=10)
progress_frame.pack(pady=5, fill='x')

progress_label = ttk.Label(progress_frame, text="")
progress_label.pack(anchor='w')

progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(fill='x')

# Uploaded videos treeview
uploaded_frame = ttk.Frame(root, padding=10)
uploaded_frame.pack(pady=5, fill='both', expand=True)

ttk.Label(uploaded_frame, text="Uploaded Videos").pack(anchor='w')

tree_container = ttk.Frame(uploaded_frame)
tree_container.pack(fill='both', expand=True)

tree_scroll = ttk.Scrollbar(tree_container)
tree = ttk.Treeview(tree_container, columns=("Title", "URL", "Embed Code"), show="headings", yscrollcommand=tree_scroll.set)
tree.heading("Title", text="Video Title")
tree.heading("URL", text="URL")
tree.heading("Embed Code", text="Embed Code")
tree.column("Title", width=200)
tree.column("URL", width=300)
tree.column("Embed Code", width=250)
tree.pack(side='left', fill='both', expand=True)
tree_scroll.pack(side='right', fill='y')

root.protocol("WM_DELETE_WINDOW", lambda: [save_config(), root.destroy()])
root.mainloop()